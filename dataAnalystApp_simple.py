import re
import concurrent.futures
import os
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.dimensions import ColumnDimension
from io import BytesIO

import pandas as pd
import streamlit as st
import markdown
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.backends import default_backend
import plotly.io as pio
import base64
import snowflake.connector
from openai import OpenAI

client = OpenAI(api_key=st.secrets.openai_credentials.key)
st.set_page_config(page_title="AI Data Analyst", page_icon=":sparkles:", layout="wide")

pd.set_option('display.max_columns', 500)
pd.set_option('display.max_rows', 500)
pd.set_option('display.width', 1000)

# Set to True to use a custom model deployment. False to use a DataRobot Playground model.
CUSTOM_MODEL_MODE = False

def initialize_session_state():
    default_values = {
        # 'password': password,
        'businessQuestion': '',
        'askButton': False,
        'clearButton': False,
        'dictionary': '',
        'dictionary_chunks': '',
        'this_table_dictionary': '',
        'llm_generated_dictionary': '',
        'table_selection_button': False,
        'selectedTables': [],
        'selectedCSVFile': None,
        'csv_selection_button': False,
        'cache_cleared': False,
        'tables': [],
        'df': pd.DataFrame(),
        'prompt': '',
        'sqlCode': '',
        'results': pd.DataFrame(),
        'fig1': None,
        'fig2': None,
        'analysis': '',
        'suggestedQuestions': '',
        'tableDescriptions': [],
        'tableSamples': [],
        'smallTableSamples': [],
        'frequentValues': pd.DataFrame(),
        'datarobot_logo_svg': '',
        'customer_logo_svg': '',
        'html_content': '',
        'download_link': '',
        'csvUploadButton': None,
        'excel_content': None,
        'chart_code' : None
    }
    for key, value in default_values.items():
        st.session_state.setdefault(key, value)

initialize_session_state()

@st.cache_data(show_spinner=False)
def suggestQuestion(description):
    # description = "this is a test."
    systemPrompt = st.secrets.prompts.suggest_a_question
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [description]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(description)]})
    deployment_id = st.secrets.datarobot_deployment_id.summarize_table
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY

    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    suggestion = predictions_response.json()["data"][0]["prediction"]
    return suggestion


@st.cache_data(show_spinner=False)
def getDataDictionary(prompt):
    systemPrompt = st.secrets.prompts.get_data_dictionary
    # prompt = data
    # prompt = "this is a test. are you there?"

    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(prompt)]})
    deployment_id = st.secrets.datarobot_deployment_id.data_dictionary_maker
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    dictionary = predictions_response.json()["data"][0]["prediction"]
    return dictionary


@st.cache_data(show_spinner=False)
def getPythonCode(prompt):
    systemPrompt = st.secrets.prompts.get_python_code
    # prompt = "test"
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + prompt]})
    deployment_id = st.secrets.datarobot_deployment_id.python_code_generator
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    code = predictions_response.json()["data"][0]["prediction"]

    # Pattern to match code blocks that optionally start with ```python or just ```
    pattern = r'```(?:python)?\n(.*?)```'
    matches = re.findall(pattern, code, re.DOTALL)
    # st.write(matches)
    # Join all matches into a single string, separated by two newlines
    python_code = '\n\n'.join(matches)
    return python_code
def executePythonCode(prompt, df):
    '''
    Executes the Python Code generated by the LLM
    '''
    print("Generating code...")
    pythonCode = getPythonCode(prompt)
    print("Executing...")
    try:
        function_dict = {}
        exec(pythonCode, function_dict)  # execute the code created by our LLM
        analyze_data = function_dict['analyze_data']  # get the function that our code created
        results = analyze_data(df)
    except Exception as e:
        print(e)
    return pythonCode, results


@st.cache_data(show_spinner=False)
def getChartCode(prompt):
    systemPrompt = st.secrets.prompts.get_chart_code
    # prompt = "test"
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + prompt]})
    deployment_id = st.secrets.datarobot_deployment_id.plotly_code_generator
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    code = predictions_response.json()["data"][0]["prediction"]
    # Pattern to match code blocks that optionally start with ```python or just ```
    pattern = r'```(?:python)?\n(.*?)```'
    matches = re.findall(pattern, code, re.DOTALL)

    # Join all matches into a single string, separated by two newlines
    chart_code = '\n\n'.join(matches)
    return chart_code
@st.cache_data(show_spinner=False)
def createCharts(prompt, results):
    print("getting chart code...")
    chartCode = getChartCode(prompt + str(results))
    st.session_state["chart_code"] = chartCode
    print(chartCode.replace("```python", "").replace("```", ""))
    function_dict = {}
    exec(chartCode.replace("```python", "").replace("```", ""), function_dict)  # execute the code created by our LLM
    print("executing chart code...")
    create_charts = function_dict['create_charts']  # get the function that our code created
    fig1, fig2 = create_charts(results)
    return fig1, fig2
@st.cache_data(show_spinner=False)
def getBusinessAnalysis(prompt):
    systemPrompt = st.secrets.prompts.get_business_analysis
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + prompt]})
    deployment_id = st.secrets.datarobot_deployment_id.business_analysis
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    business_analysis = predictions_response.json()["data"][0]["prediction"]
    return business_analysis
@st.cache_data(show_spinner=False)
def get_top_frequent_values(df):
    # Select non-numeric columns
    non_numeric_cols = df.select_dtypes(exclude=['number']).columns

    # Prepare a list to store the results
    results = []

    # Iterate over non-numeric columns
    for col in non_numeric_cols:
        # Find top 10 most frequent values for the column
        top_values = df[col].value_counts().head(10).index.tolist()

        # Convert the values to strings
        top_values = [str(value) for value in top_values]

        # Append the column name and its frequent values to the results
        results.append({'Non-numeric column name': col, 'Frequent Values': top_values})

    # Create a new DataFrame for the results
    result_df = pd.DataFrame(results)

    return result_df

# Function that creates the charts and business analysis
@st.cache_data(show_spinner=False)
def createChartsAndBusinessAnalysis(businessQuestion, results, prompt):
    attempt_count = 0
    max_attempts = 6
    fig1 = fig2 = None
    analysis = None

    with concurrent.futures.ThreadPoolExecutor() as executor:
        while attempt_count < max_attempts:
            chart_future = executor.submit(createCharts, businessQuestion, results)
            analysis_future = executor.submit(getBusinessAnalysis, prompt + str(results))
            try:
                if fig1 is None or fig2 is None:
                    fig1, fig2 = chart_future.result(timeout=30)  # Add a timeout for better handling
                    with st.expander(label="Charts", expanded=True):
                        st.plotly_chart(fig1, theme="streamlit", use_container_width=True)
                        st.plotly_chart(fig2, theme="streamlit", use_container_width=True)
                break  # If operation succeeds, break out of the loop
            except Exception as e:
                attempt_count += 1
                print(f"Chart Attempt {attempt_count} failed with error: {repr(e)}")
                fig1_str = str(fig1) if fig1 is not None else "None"
                fig2_str = str(fig2) if fig2 is not None else "None"
                businessQuestion += f"\nCHART CODE FAILED!  Attempt {attempt_count} failed with error: {repr(e)}\nFig1: {fig1_str}\nFig2: {fig2_str}"

                if attempt_count >= max_attempts:
                    print("Max charting attempts reached, handling the failure.")
                    st.write("I was unable to plot the data.")
                    # Handle the failure after the final attempt
                else:
                    print("Retrying the charts...")

        try:
            with st.expander(label="Business Analysis", expanded=True):
                analysis = analysis_future.result(timeout=30)  # Add a timeout for better handling
                st.markdown(analysis.replace("$", "\$"))
        except:
            st.write("I am unable to provide the analysis. Please rephrase the question and try again.")

    return fig1, fig2, analysis

# Function to create a download link
@st.cache_data(show_spinner=False)
def create_download_link(html_content, filename):
    b64 = base64.b64encode(html_content.encode()).decode()  # B64 encode
    href = f'<a href="data:text/html;base64,{b64}" download="{filename}">Download as HTML</a>'
    return href

@st.cache_data(show_spinner=False)
def read_svg(file_path):
    with open(file_path, 'r') as file:
        content = file.read()
    return content
@st.cache_data(show_spinner=False)
def read_svg_as_base64(file_path):
    with open(file_path, 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')

# Callback function to generate HTML content
@st.cache_data(show_spinner=False)
def generate_html_report(businessQuestion, sqlcode, results, fig1, fig2, analysis, datarobot_logo_svg, customer_logo_svg):
    plotly_html1 = pio.to_html(fig1, full_html=False, include_plotlyjs=True, default_width="100%",
                               default_height="100%")
    plotly_html2 = pio.to_html(fig2, full_html=False, include_plotlyjs=True, default_width="100%",
                               default_height="100%")

    # Convert markdown to HTML for the analysis section
    if analysis and analysis.strip():
        analysis_html = markdown.markdown(analysis)
    else:
        st.error("No analysis data found to generate the report.")

    html_content = f"""
    <html>
    <head>
        <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;700&family=IBM+Plex+Mono:wght@400;700&display=swap">
        <style>
            body {{
                font-family: 'IBM Plex Sans', sans-serif;
                background-color: #F0F1F2;
                color: #0D0D0D;
                margin: 40px;
            }}
            h1, h2 {{
                font-family: 'IBM Plex Sans', sans-serif;
                color: #0D0D0D;
            }}
            pre, code {{
                font-family: 'IBM Plex Mono', monospace;
            }}
            .report-title {{
                font-size: 2.5em;
                font-weight: bold;
                text-align: left;
                margin-top: 40px;
            }}
            .section-title {{
                font-size: 1.75em;
                font-weight: bold;
                margin-top: 20px;
            }}
            .logo-container {{
                text-align: left;
                margin-bottom: 20px;
            }}
            .logo-datarobot {{
                width: 300px;
                margin-bottom: 10px;
                display: block;
            }}
            .logo-customer {{
                width: 300px;
                margin-bottom: 10px;
                display: block;
            }}
            .horizontal-rule {{
                border: 0;
                height: 2px;
                background: #03A688;
                margin: 20px 0;
            }}
            .collapsible {{
                background-color: #03A688;
                color: white;
                cursor: pointer;
                padding: 10px;
                width: 100%;
                border: none;
                text-align: left;
                outline: none;
                font-size: 18px;
                font-weight: bold;
            }}
            .collapsible:after {{
                content: '+';
                font-size: 18px;
                float: right;
            }}
            .collapsible.active:after {{
                content: '-';
            }}
            .content {{
                padding: 0 18px;
                display: none;
                overflow: hidden;
                background-color: #f9f9f9;
            }}
            .content.show {{
                display: block;
            }}
        </style>
        <title>AI Data Analyst Report</title>
    </head>
    <body>
        <div class="logo-container">            
            <img src="data:image/svg+xml;base64,{datarobot_logo_svg}" class="logo-customer" alt="Customer Logo">
        </div>
        <h1 class="report-title">AI Data Analyst Report</h1>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible active">Business Question</button>
        <div class="content show">
            <p>{businessQuestion}</p>
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible">Analysis Code</button>
        <div class="content">
            <pre>{sqlcode}</pre>
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible">Results</button>
        <div class="content">
            {results.to_html(index=False, escape=False)}
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible active">Charts</button>
        <div class="content show">
            <div>{plotly_html1}</div>
            <div>{plotly_html2}</div>
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible active">Business Analysis</button>
        <div class="content show">
            <div>{analysis_html}</div>
        </div>
        <script>
            var coll = document.getElementsByClassName("collapsible");
            for (var i = 0; i < coll.length; i++) {{
                coll[i].addEventListener("click", function() {{
                    this.classList.toggle("active");
                    var content = this.nextElementSibling;
                    if (content.style.display === "block" || content.classList.contains("show")) {{
                        content.style.display = "none";
                        content.classList.remove("show");
                    }} else {{
                        content.style.display = "block";
                        content.classList.add("show");
                    }}
                }});
            }}
        </script>
    </body>
    </html>
    """
    return html_content

@st.cache_data(show_spinner=False)
def process_tables(dictionary, selectedTables, sampleSize):
    tableSamples = []
    tableDescriptions = []
    frequentValues = pd.DataFrame()

    for table in selectedTables:
        tableDescription = summarizeTable(dictionary, table)
        results = getTableSample(sampleSize=sampleSize, table=table)
        tableSamples.append(results)
        tableDescriptions.append(tableDescription)
        freqVals = get_top_frequent_values(results)
        frequentValues = pd.concat([frequentValues, freqVals], axis=0)

    smallTableSamples = []
    for table in tableSamples:
        smallSample = table.sample(n=3)
        smallTableSamples.append(smallSample)

    return tableDescriptions, tableSamples, smallTableSamples, frequentValues


def text_input_enterKey():
    st.session_state["askButton"] = True

def clear_text():
    st.session_state["businessQuestion"] = ""
    st.session_state["askButton"] = False

def make_dictionary_chunks(df):
    dictionary_chunks = []
    chunk_size = 15
    total_columns = len(df.columns)
    progress_placeholder = st.empty()

    for start in range(0, total_columns, chunk_size):
        current_chunk = start // chunk_size + 1
        total_chunks = (total_columns + chunk_size - 1) // chunk_size
        progress = current_chunk / total_chunks

        with progress_placeholder.container():
            st.progress(progress,
                        text=f'Processing {chunk_size} columns at a time in chunks. Currently working on chunk {current_chunk} of {total_chunks}')

        end = min(start + chunk_size, total_columns)
        subset = df.iloc[:10, start:end]
        data = "First 10 Rows: \n" + str(
            subset) + "\n Unique and Frequent Values of Categorical Data: \n" + str(
            get_top_frequent_values(df))

        dictionary_chunk = getDataDictionary(data)
        dictionary_chunks.append(dictionary_chunk)

    progress_placeholder.empty()
    return dictionary_chunks

@st.cache_data(show_spinner=False)
def assembleDictionaryParts(parts):
    systemPrompt = st.secrets.prompts.assemble_data_dictionary
    # parts = data

    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [parts]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(parts)]})
    deployment_id = st.secrets.datarobot_deployment_id.data_dictionary_assembler
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    assembled = predictions_response.json()["data"][0]["prediction"]
    return assembled

def setup_sidebar():
    with st.sidebar:
        st.image("logo.png", width=300)
        st.write("Welcome to InstaData, your virtual data analyst!")

        # st.image("csv_File_Logo.svg", width=45)
        st.session_state["csvUploadButton"] = st.file_uploader(label="Upload a CSV file",
                                                               accept_multiple_files=False)
        process_csv_upload()

        with st.expander("Clear Cache", expanded=False):
            st.write("To reset any saved data and completely start over, clear the cache. You will have to reload your dataset.")
            st.button("Clear Cache", on_click=clear_cache_callback)
            if st.session_state["cache_cleared"]:
                st.success("Cache cleared successfully!")
                # Reset the flag
                st.session_state["cache_cleared"] = False


def process_csv_upload():
    if st.session_state["csvUploadButton"] is not None:
        st.session_state["selectedCSVFile"] = st.session_state["csvUploadButton"]


def display_analysis_tab(tab):
    with tab:
        st.write(st.session_state["suggestedQuestions"])

        st.session_state["businessQuestion"] = st.text_input(
            label="Question",
            # value=st.session_state["businessQuestion"],
            on_change=text_input_enterKey
        )
        display_action_buttons()

        if st.session_state.get("askButton", False):
            analyze_question()


def display_csv_explore_tab(tab):
    with tab:
        try:
            st.session_state["df"] = pd.read_csv(st.session_state["selectedCSVFile"], encoding='utf-8')
        except:
            st.session_state["df"] = pd.read_csv(st.session_state["selectedCSVFile"], encoding='ISO-8859-1')

        with st.expander(label="First 10 Rows", expanded=False):
            st.dataframe(st.session_state["df"].head(10))

        try:
            with st.expander(label="Column Descriptions", expanded=False):
                st.dataframe(st.session_state["df"].describe(include='all'))
        except:
            pass

        try:
            with st.expander(label="Unique and Frequent Values", expanded=False):
                st.dataframe(get_top_frequent_values(st.session_state["df"]))
        except Exception as e:
            print(e)

        # this section has the code to produce the contents above
        with st.expander(label="code", expanded=False):
            code = '''
                import pandas as pd
                import matplotlib.pyplot as plt
                import numpy as np

                # read the data
                try:
                    df = pd.read_csv("selectedCSVFile", encoding = 'ISO-8859-1')
                except:
                    df = pd.read_csv("selectedCSVFile")

                # display the first few rows in the data
                df.head(10)
                
                # Descriptive statistics for numeric variables
                df.describe()
                
                # Display the most frequent values for each categorical variable
                def get_top_frequent_values(df):
                    # Select non-numeric columns
                    non_numeric_cols = df.select_dtypes(exclude=['number']).columns
                
                    # Prepare a list to store the results
                    results = []
                
                    # Iterate over non-numeric columns
                    for col in non_numeric_cols:
                        # Find top 10 most frequent values for the column
                        top_values = df[col].value_counts().head(10).index.tolist()
                
                        # Convert the values to strings
                        top_values = [str(value) for value in top_values]
                
                        # Append the column name and its frequent values to the results
                        results.append({'Non-numeric column name': col, 'Frequent Values': top_values})
                
                    # Create a new DataFrame for the results
                    result_df = pd.DataFrame(results)
                    return result_df
                result_df
                '''
            st.code(code, language="python")

        try:
            with st.expander(label="Data Dictionary", expanded=True):
                with st.spinner("Making dictionary..."):
                    st.session_state['dictionary_chunks'] = make_dictionary_chunks(st.session_state["df"])
                with st.spinner("Putting it all together..."):
                    st.session_state["dictionary"] = assembleDictionaryParts(st.session_state['dictionary_chunks'])
                    st.markdown(st.session_state["dictionary"])
        except:
            pass

def display_csv_analysis_tab(tab):
    with tab:
        st.session_state["suggestedQuestions"] = suggestQuestion(st.session_state["dictionary"])
        st.write(st.session_state["suggestedQuestions"])

        st.session_state["businessQuestion"] = st.text_input(
            label="Question",
            # value=st.session_state["businessQuestion"],
            on_change=text_input_enterKey
        )
        display_action_buttons()

        if st.session_state.get("askButton", False):
            analyze_question_csv()

def display_action_buttons():
    buttonContainer = st.container()
    buttonCol1, buttonCol2, _ = buttonContainer.columns([1, 1, 8])

    buttonCol1.button(label="Ask", use_container_width=True, type="primary", on_click=text_input_enterKey)
    buttonCol2.button(label="clear", use_container_width=True, type="secondary", on_click=clear_text)


def analyze_question_csv():
    with st.spinner("Analyzing... "):
        st.session_state["prompt"] = generate_csv_prompt()
        execute_query_with_retries(csv_mode=True)

        try:
            display_query_results()
        except:
            st.write(
                "I tried a few different ways, but couldn't get a working solution. Rephrase the question and try again.")

        if st.session_state["results"] is not None and not st.session_state["results"].empty:
            analyze_and_generate_report_csv()
        else:
            st.write("The query returns an empty result. Try rephrasing the question.")
            print("No data returned.")
            st.stop()


def generate_csv_prompt():
    return ("Business Question: " + str(st.session_state["businessQuestion"]) +
            "\n Data Sample: \n" + str(st.session_state["df"].head(3)) +
            "\n Unique and Frequent Values of Categorical Data: \n" + str(
                get_top_frequent_values(st.session_state["df"])) +
            "\n Data Dictionary: \n" + str(st.session_state["dictionary"]))

def execute_query_with_retries(csv_mode):
    attempts = 0
    max_retries = 5
    while attempts < max_retries:
        st.session_state["sqlCode"] = None
        try:
            if csv_mode:
                st.session_state["sqlCode"], st.session_state["results"] = executePythonCode(st.session_state["prompt"], st.session_state["df"])
            else:
                st.session_state["sqlCode"], st.session_state["results"] = executeSnowflakeQuery(st.session_state["prompt"], user, st.session_state["password"], account, warehouse, database, schema)
                # st.session_state["sqlCode"], st.session_state["results"] = executeSnowflakeSnowpark(st.session_state["prompt"], user, st.session_state["password"], account, warehouse, database, schema)
            if st.session_state["results"].empty:
                raise ValueError("The DataFrame is empty, retrying...")
            break
        except Exception as e:
            attempts += 1
            st.session_state[
                "prompt"] += f"\nQUERY FAILED! Attempt {attempts} failed with error: {repr(e)}\nCode: {st.session_state['sqlCode']}"
            if attempts == max_retries:
                break

def display_query_results():
    with st.expander(label="Code", expanded=False):
        st.code(st.session_state["sqlCode"], language="sql")
    with st.expander(label="Result", expanded=False):
        st.table(st.session_state["results"])

def analyze_and_generate_report(full_dictionary):
    with st.spinner("Visualization and analysis in progress..."):
        st.session_state["fig1"], st.session_state["fig2"], st.session_state[
            "analysis"] = createChartsAndBusinessAnalysis(
            st.session_state["businessQuestion"],
            st.session_state["results"], st.session_state["prompt"])

    generate_report(full_dictionary)

def analyze_and_generate_report_csv():
    with st.spinner("Visualization and analysis in progress..."):
        st.session_state["fig1"], st.session_state["fig2"], st.session_state[
            "analysis"] = createChartsAndBusinessAnalysis(
            st.session_state["businessQuestion"],
            st.session_state["results"], st.session_state["prompt"])

    generate_report_csv()

def generate_report(full_dictionary):
    read_svgs_and_generate_html_report()
    create_and_display_download_link()
    read_svgs_and_generate_excel_report()
    create_and_display_excel_download_link()

def generate_report_csv():
    read_svgs_and_generate_html_report()
    create_and_display_download_link()
    read_svgs_and_generate_excel_report()
    create_and_display_excel_download_link()


def read_svgs_and_generate_html_report():
    st.session_state["datarobot_logo_svg"] = read_svg_as_base64("DataRobotLogo.svg")
    st.session_state["customer_logo_svg"] = read_svg_as_base64("small_square_placeholder.svg")

    st.session_state["html_content"] = generate_html_report(st.session_state["businessQuestion"],
                                                            st.session_state["sqlCode"],
                                                            st.session_state["results"], st.session_state["fig1"],
                                                            st.session_state["fig2"],
                                                            st.session_state["analysis"],
                                                            st.session_state["datarobot_logo_svg"],
                                                            st.session_state["customer_logo_svg"])
def create_and_display_download_link():
    try:
        st.session_state["download_link"] = create_download_link(st.session_state["html_content"], 'report.html')
        st.markdown(st.session_state["download_link"], unsafe_allow_html=True)
    except:
        pass
@st.cache_data(show_spinner=False)
def create_download_link_excel(excel_data, filename):
    if not excel_data:
        st.error("Excel content is empty. Cannot create a download link.")
        return ""
    b64 = base64.b64encode(excel_data).decode()  # B64 encode
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download as Excel</a>'
    return href

@st.cache_data(show_spinner=False)
def read_svg(file_path):
    with open(file_path, 'r') as file:
        content = file.read()
    return content

@st.cache_data(show_spinner=False)
def read_svg_as_base64(file_path):
    with open(file_path, 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')

# Callback function to generate Excel content
@st.cache_data(show_spinner=False)
def generate_excel_report(businessQuestion, sqlcode, results, fig1, fig2, analysis):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    try:
        # Writing data to Excel
        if businessQuestion:
            df_business_question = pd.DataFrame({'Business Question': [businessQuestion]})
            df_business_question.to_excel(writer, index=False, sheet_name='Business Question')
            worksheet = writer.sheets['Business Question']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        if results is not None and not results.empty:
            results.to_excel(writer, index=False, sheet_name='Results')
            worksheet = writer.sheets['Results']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
            worksheet.auto_filter.ref = worksheet.dimensions

        # Writing analysis to Excel
        if analysis and analysis.strip():
            analysis_sections = {
                "The Bottom Line": "",
                "Additional Insights": "",
                "Follow Up Questions": ""
            }
            current_section = None
            for line in analysis.splitlines():
                line = line.strip()
                if line.startswith("###"):
                    header = line.replace("###", "").strip()
                    if header in analysis_sections:
                        current_section = header
                elif current_section:
                    analysis_sections[current_section] += line + "\n"

            # Write each section to individual cells
            worksheet = writer.book.create_sheet(title='Analysis')
            row = 1
            for section, content in analysis_sections.items():
                worksheet[f'A{row}'] = section
                worksheet[f'A{row}'].font = Font(bold=True)
                cell = worksheet[f'A{row + 1}']
                # content = content.strip().replace('', '')  # Replace newlines for better formatting
                cell.value = content
                cell.alignment = Alignment(wrap_text=True)
                worksheet.column_dimensions['A'].width = 50
                row += 3

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        else:
            st.error("No analysis data found to generate the report.")

        # Add Plotly charts as images
        worksheet = writer.book.create_sheet(title="Charts")
        if fig1:
            fig1_bytes = fig1.to_image(format="png")
            img1 = Image(BytesIO(fig1_bytes))
            worksheet.add_image(img1, "A1")
        if fig2:
            fig2_bytes = fig2.to_image(format="png")
            img2 = Image(BytesIO(fig2_bytes))
            worksheet.add_image(img2, "A20")

        writer._save()  # Close the writer before accessing the value
        return output.getvalue()
    except Exception as e:
        st.error(f"An error occurred while generating the Excel report: {e}")
        return None

def read_svgs_and_generate_excel_report():
    st.session_state["datarobot_logo_svg"] = read_svg_as_base64("DataRobotLogo.svg")
    st.session_state["customer_logo_svg"] = read_svg_as_base64("small_square_placeholder.svg")

    st.session_state["excel_content"] = generate_excel_report(st.session_state.get("businessQuestion"),
                                                               st.session_state.get("sqlCode"),
                                                               st.session_state.get("results"), st.session_state.get("fig1"),
                                                               st.session_state.get("fig2"),
                                                               st.session_state.get("analysis"))

def create_and_display_excel_download_link():
    st.session_state["download_link_excel"] = create_download_link_excel(st.session_state.get("excel_content"), 'report.xlsx')
    if st.session_state["download_link_excel"]:
        st.markdown(st.session_state["download_link_excel"], unsafe_allow_html=True)




def clear_cache_callback():
    # Clear both data and resource caches
    st.cache_data.clear()
    st.cache_resource.clear()

    # Update session state to show success message
    st.session_state["cache_cleared"] = True

def mainPage():
    setup_sidebar()

    st.markdown("InstaData helps you quickly get insights from your data!")
    st.header("Upload a CSV file on the left to get started!")

    tab1, tab2 = st.tabs(["Analyze", "Explore"])
    if st.session_state["selectedCSVFile"]:
    #     tab1, tab2 = st.tabs(["Analyze", "Explore"])

        with st.spinner(text="Analyzing table structure, see Explore tab for details..."):
            display_csv_explore_tab(tab2)
        display_csv_analysis_tab(tab1)

# Main app
def _main():
    hide_streamlit_style = """
    <style>
    # MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """
    # st.markdown(hide_streamlit_style, unsafe_allow_html=True)  # This lets you hide the Streamlit branding

    mainPage()


if __name__ == "__main__":
    _main()
