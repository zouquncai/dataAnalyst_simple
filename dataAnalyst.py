import re
import concurrent.futures
import os
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.dimensions import ColumnDimension
from io import BytesIO

import pandas as pd
import streamlit as st
import markdown
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.backends import default_backend
import plotly.io as pio
import base64
import snowflake.connector
from openai import OpenAI

client = OpenAI(api_key=st.secrets.openai_credentials.key)
st.set_page_config(page_title="AI Data Analyst", page_icon=":sparkles:", layout="wide")

pd.set_option('display.max_columns', 500)
pd.set_option('display.max_rows', 500)
pd.set_option('display.width', 1000)

# Set to True to use a custom model deployment. False to use a DataRobot Playground model.
CUSTOM_MODEL_MODE = False

# Snowflake connection details
user = st.secrets.snowflake_credentials.user
password = st.secrets.snowflake_credentials.password
account = st.secrets.snowflake_credentials.account
warehouse = st.secrets.snowflake_credentials.warehouse
database = st.secrets.snowflake_credentials.database
schema = st.secrets.snowflake_credentials.schema

def initialize_session_state():
    default_values = {
        'password': password,
        'businessQuestion': '',
        'askButton': False,
        'clearButton': False,
        'dictionary': '',
        'dictionary_chunks': '',
        'this_table_dictionary': '',
        'llm_generated_dictionary': '',
        'snowflake_submit_button': False,
        'table_selection_button': False,
        'selectedTables': [],
        'selectedCSVFile': None,
        'csv_selection_button': False,
        'cache_cleared': False,
        'tables': [],
        'df': pd.DataFrame(),
        'prompt': '',
        'sqlCode': '',
        'results': pd.DataFrame(),
        'fig1': None,
        'fig2': None,
        'analysis': '',
        'suggestedQuestions': '',
        'tableDescriptions': [],
        'tableSamples': [],
        'smallTableSamples': [],
        'frequentValues': pd.DataFrame(),
        'datarobot_logo_svg': '',
        'customer_logo_svg': '',
        'html_content': '',
        'download_link': '',
        'csvUploadButton': None,
        'excel_content': None,
        'chart_code' : None
    }
    for key, value in default_values.items():
        st.session_state.setdefault(key, value)

initialize_session_state()
@st.cache_data(show_spinner=False)
def getSnowflakeTableDescriptions(tables, user, password, account, warehouse, database, schema):
    # Establish a connection to Snowflake
    try:
        conn = snowflake.connector.connect(
            user=user,
            password=password,
            account=account,
            warehouse=warehouse,
            database=database,
            schema=schema,
            # Enable case sensitivity for identifiers
            case_sensitive_identifier_quoting=True
        )
        cursor = conn.cursor()
    except Exception as e:
        print(f"Error connecting to Snowflake: {e}")
        return None

    # Function to get primary keys of a table
    def get_primary_keys(table_name):
        try:
            cursor.execute(f"""
                SELECT COLUMN_NAME
                FROM {database}.INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
                JOIN {database}.INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu
                ON tc.CONSTRAINT_NAME = kcu.CONSTRAINT_NAME
                WHERE tc.CONSTRAINT_TYPE = 'PRIMARY KEY'
                AND tc.TABLE_SCHEMA = '{schema}'
                AND tc.TABLE_NAME = '{table_name}'
                """)
            return [row[0] for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error fetching primary keys for table {table_name}: {e}")
            return []

    # Function to get columns and data types along with additional metadata
    def get_columns_and_types(table_name):
        try:
            cursor.execute(f"""
                SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT, COMMENT
                FROM {database}.INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = '{schema}'
                AND TABLE_NAME = '{table_name}'
                """)
            columns = cursor.fetchall()
            primary_keys = get_primary_keys(table_name)
            return [(col[0], col[1], col[2] == 'YES', col[3], col[0] in primary_keys, col[4]) for col in columns]
        except Exception as e:
            print(f"Error fetching columns and types for table {table_name}: {e}")
            return []

    # Function to get table comment
    def get_table_comment(table_name):
        try:
            cursor.execute(f"""
                SELECT COMMENT
                FROM {database}.INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = '{schema}'
                AND TABLE_NAME = '{table_name}'
                """)
            result = cursor.fetchone()
            return result[0] if result else None
        except Exception as e:
            print(f"Error fetching table comment for {table_name}: {e}")
            return None

    # Function to get table row count
    def get_table_row_count(table_name):
        try:
            cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name}")
            result = cursor.fetchone()
            return result[0] if result else None
        except Exception as e:
            print(f"Error fetching row count for table {table_name}: {e}")
            return None

    # Prepare the descriptions string
    descriptions = ""

    for table in tables:
        descriptions += f"Table: {table}\n"
        table_comment = get_table_comment(table)
        if table_comment:
            descriptions += f" Comment: {table_comment}\n"
        row_count = get_table_row_count(table)
        descriptions += f" Row Count: {row_count}\n"
        for col_name, col_type, nullable, default, is_primary, col_comment in get_columns_and_types(table):
            descriptions += f' Column: "{col_name}", Type: {col_type}, Nullable: {nullable}, Default: {default}, Primary Key: {is_primary}, Comment: {col_comment}\n'
        descriptions += "---------------------------------------------------------------\n"

    # Close the connection
    cursor.close()
    conn.close()

    return descriptions


@st.cache_data(show_spinner=False)
def suggestQuestion(description):
    # description = "this is a test."
    systemPrompt = st.secrets.prompts.suggest_a_question
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [description]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(description)]})
    deployment_id = st.secrets.datarobot_deployment_id.summarize_table
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY

    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    suggestion = predictions_response.json()["data"][0]["prediction"]
    return suggestion

@st.cache_data(show_spinner=False)
def summarizeTable(dictionary, table):
    systemPrompt = st.secrets.prompts.summarize_table
    systemPrompt = systemPrompt.format(table=table)
    # table = "This is a test"
    # dictionary = "this is a test dictionary."
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [str(dictionary) + "\nTABLE TO DESCRIBE: " + str(table)]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(dictionary) + "\nTABLE TO DESCRIBE: " + str(table)]})
    deployment_id = st.secrets.datarobot_deployment_id.summarize_table
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    summary = predictions_response.json()["data"][0]["prediction"]
    return summary

@st.cache_data(show_spinner=False)
def getDataDictionary(prompt):
    systemPrompt = st.secrets.prompts.get_data_dictionary
    # prompt = data
    # prompt = "this is a test. are you there?"

    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(prompt)]})
    deployment_id = st.secrets.datarobot_deployment_id.data_dictionary_maker
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    dictionary = predictions_response.json()["data"][0]["prediction"]
    return dictionary

@st.cache_data(show_spinner=False)
def assembleDictionaryParts(parts):
    systemPrompt = st.secrets.prompts.assemble_data_dictionary
    # parts = data

    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [parts]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + str(parts)]})
    deployment_id = st.secrets.datarobot_deployment_id.data_dictionary_assembler
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    assembled = predictions_response.json()["data"][0]["prediction"]
    return assembled
@st.cache_data(show_spinner=False)
def getPythonCode(prompt):
    systemPrompt = st.secrets.prompts.get_python_code
    # prompt = "test"
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + prompt]})
    deployment_id = st.secrets.datarobot_deployment_id.python_code_generator
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    code = predictions_response.json()["data"][0]["prediction"]
    # Pattern to match code blocks that optionally start with ```python or just ```
    pattern = r'```(?:python)?\n(.*?)```'
    matches = re.findall(pattern, code, re.DOTALL)

    # Join all matches into a single string, separated by two newlines
    python_code = '\n\n'.join(matches)
    return python_code
def executePythonCode(prompt, df):
    '''
    Executes the Python Code generated by the LLM
    '''
    print("Generating code...")
    pythonCode = getPythonCode(prompt)
    print("Executing...")
    try:
        function_dict = {}
        exec(pythonCode, function_dict)  # execute the code created by our LLM
        analyze_data = function_dict['analyze_data']  # get the function that our code created
        results = analyze_data(df)
    except Exception as e:
        print(e)
    return pythonCode, results
@st.cache_data(show_spinner=False)
def getSnowflakeSQL(prompt, warehouse=warehouse, database=database, schema=schema):
    systemPrompt = st.secrets.prompts.get_snowflake_sql
    systemPrompt = systemPrompt.format(warehouse=warehouse, database=database, schema=schema)
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [
            str(prompt) + "\nSNOWFLAKE ENVIRONMENT:\nwarehouse = " + str(warehouse) + "\ndatabase = " + str(
                database) + "\nschema = " + str(schema)]})
    else:
        data = pd.DataFrame({"promptText": [
            systemPrompt + "\n\n" + str(prompt) + "\nSNOWFLAKE ENVIRONMENT:\nwarehouse = " + str(
                warehouse) + "\ndatabase = " + str(database) + "\nschema = " + str(schema)]})
    deployment_id = st.secrets.datarobot_deployment_id.sql_code_generator
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    st.write(url)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    code = predictions_response.json()["data"][0]["prediction"]
    # Pattern to match code blocks that optionally start with ```python or just ```
    pattern = r'```(?:sql)?\n(.*?)```'
    matches = re.findall(pattern, code, re.DOTALL)

    # Join all matches into a single string, separated by two newlines
    sql_code = '\n\n'.join(matches)
    return sql_code
@st.cache_data(show_spinner=False)
def executeSnowflakeQuery(prompt, user, password, account, warehouse, database, schema):
    # prompt = "Retrieve a random sample using SAMPLE(1000 ROWS) from this table: 'LENDING_CLUB_TARGET'"
    # Get the SQL code
    snowflakeSQL = getSnowflakeSQL(prompt)

    # Create a connection using Snowflake Connector
    conn = snowflake.connector.connect(
        user=user,
        password=password,
        account=account,
        warehouse=warehouse,
        database=database,
        schema=schema,
        # Enable case sensitivity for identifiers
        case_sensitive_identifier_quoting=True
    )
    results = None

    try:
        # Execute the query and fetch the results into a DataFrame
        with conn.cursor() as cur:
            cur.execute(snowflakeSQL)
            results = cur.fetch_pandas_all()
            # results.columns = results.columns.str.upper()
    except snowflake.connector.errors.Error as e:
        print(f"An error occurred: {e}")
    finally:
        conn.close()

    return snowflakeSQL, results
@st.cache_data(show_spinner=False)
def getSnowflakePython(prompt, warehouse=warehouse, database=database, schema=schema):
    systemPrompt = st.secrets.prompts.get_snowflake_snowpark
    systemPrompt = systemPrompt.format(warehouse=warehouse, database=database, schema=schema)
    data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [
        str(prompt) + "\nSNOWFLAKE ENVIRONMENT:\nwarehouse = " + str(warehouse) + "\ndatabase = " + str(
            database) + "\nschema = " + str(schema)]})
    deployment_id = st.secrets.datarobot_deployment_id.sql_code_generator
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    code = predictions_response.json()["data"][0]["prediction"]
    # Pattern to match code blocks that optionally start with ```python or just ```
    pattern = r'```(?:python)?\n(.*?)```'
    matches = re.findall(pattern, code, re.DOTALL)

    # Join all matches into a single string, separated by two newlines
    snowpark_code = '\n\n'.join(matches)
    return snowpark_code
@st.cache_data(show_spinner=False)
def executeSnowflakeSnowpark(prompt, user, password, account, warehouse, database, schema):
    from snowflake.snowpark import Session
    import snowflake.snowpark.functions as F

    # Get the Snowpark Python DataFrame transformation as a string
    snowflake_df_transform = getSnowflakePython(prompt)

    print("SNOWPARK CODE\n================")
    print(snowflake_df_transform)

    # Define connection parameters
    connection_parameters = {
        "account": account,
        "user": user,
        "password": password,
        "warehouse": warehouse,
        "database": database,
        "schema": schema
    }

    # Create a Snowflake session
    session = Session.builder.configs(connection_parameters).create()
    results = None

    try:
        # Combine the imports and the transform function in one execution block
        exec(snowflake_df_transform, globals(), locals())

        # Assume the code defines a function called 'transform_df' that takes a session
        if 'transform_df' in locals():
            df = locals()['transform_df'](session)
        else:
            raise ValueError("The code did not define a 'transform_df' function.")

        # Convert the Snowpark DataFrame to a Pandas DataFrame
        results = df.to_pandas()
        results.columns = results.columns.str.upper()
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        session.close()

    return snowflake_df_transform, results

@st.cache_data(show_spinner=False)
def getDataSample(sampleSize):
    sampleSQLprompt = f"""
                      Select a {sampleSize} row random sample using the SAMPLE clause                
                      """
    sampleSQL = getSnowflakeSQL(sampleSQLprompt)

    sql, sample = executeSnowflakeQuery(sampleSQL, user, st.session_state["password"], account, warehouse, database,
                                        schema)
    return sample
@st.cache_data(show_spinner=False)
def getTableSample(sampleSize, table):
    sqlCode, results = executeSnowflakeQuery(
        f"Retrieve a random sample using SAMPLE({sampleSize} ROWS) from this table: " + str(table), user,
        password, account, warehouse, database, schema)
    return results
@st.cache_data(show_spinner=False)
def getChartCode(prompt):
    systemPrompt = st.secrets.prompts.get_chart_code
    # prompt = "test"
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + prompt]})
    deployment_id = st.secrets.datarobot_deployment_id.plotly_code_generator
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    code = predictions_response.json()["data"][0]["prediction"]
    # Pattern to match code blocks that optionally start with ```python or just ```
    pattern = r'```(?:python)?\n(.*?)```'
    matches = re.findall(pattern, code, re.DOTALL)

    # Join all matches into a single string, separated by two newlines
    chart_code = '\n\n'.join(matches)
    return chart_code
@st.cache_data(show_spinner=False)
def createCharts(prompt, results):
    print("getting chart code...")
    chartCode = getChartCode(prompt + str(results))
    st.session_state["chart_code"] = chartCode
    print(chartCode.replace("```python", "").replace("```", ""))
    function_dict = {}
    exec(chartCode.replace("```python", "").replace("```", ""), function_dict)  # execute the code created by our LLM
    print("executing chart code...")
    create_charts = function_dict['create_charts']  # get the function that our code created
    fig1, fig2 = create_charts(results)
    return fig1, fig2
@st.cache_data(show_spinner=False)
def getBusinessAnalysis(prompt):
    systemPrompt = st.secrets.prompts.get_business_analysis
    if CUSTOM_MODEL_MODE:
        data = pd.DataFrame({"systemPrompt": systemPrompt, "promptText": [prompt]})
    else:
        data = pd.DataFrame({"promptText": [systemPrompt + "\n\n" + prompt]})
    deployment_id = st.secrets.datarobot_deployment_id.business_analysis
    API_URL = f'{st.secrets.datarobot_credentials.PREDICTION_SERVER}/predApi/v1.0/deployments/{deployment_id}/predictions'
    API_KEY = st.secrets.datarobot_credentials.API_KEY
    DATAROBOT_KEY = st.secrets.datarobot_credentials.DATAROBOT_KEY
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': 'Bearer {}'.format(API_KEY),
        'DataRobot-Key': DATAROBOT_KEY,
    }
    url = API_URL.format(deployment_id=deployment_id)
    predictions_response = requests.post(
        url,
        data=data.to_json(orient='records'),
        headers=headers
    )
    business_analysis = predictions_response.json()["data"][0]["prediction"]
    return business_analysis
@st.cache_data(show_spinner=False)
def get_top_frequent_values(df):
    # Select non-numeric columns
    non_numeric_cols = df.select_dtypes(exclude=['number']).columns

    # Prepare a list to store the results
    results = []

    # Iterate over non-numeric columns
    for col in non_numeric_cols:
        # Find top 10 most frequent values for the column
        top_values = df[col].value_counts().head(10).index.tolist()

        # Convert the values to strings
        top_values = [str(value) for value in top_values]

        # Append the column name and its frequent values to the results
        results.append({'Non-numeric column name': col, 'Frequent Values': top_values})

    # Create a new DataFrame for the results
    result_df = pd.DataFrame(results)

    return result_df

# Function that creates the charts and business analysis
@st.cache_data(show_spinner=False)
def createChartsAndBusinessAnalysis(businessQuestion, results, prompt):
    attempt_count = 0
    max_attempts = 6
    fig1 = fig2 = None
    analysis = None

    with concurrent.futures.ThreadPoolExecutor() as executor:
        while attempt_count < max_attempts:
            chart_future = executor.submit(createCharts, businessQuestion, results)
            analysis_future = executor.submit(getBusinessAnalysis, prompt + str(results))
            try:
                if fig1 is None or fig2 is None:
                    fig1, fig2 = chart_future.result(timeout=30)  # Add a timeout for better handling
                    with st.expander(label="Charts", expanded=True):
                        st.plotly_chart(fig1, theme="streamlit", use_container_width=True)
                        st.plotly_chart(fig2, theme="streamlit", use_container_width=True)
                break  # If operation succeeds, break out of the loop
            except Exception as e:
                attempt_count += 1
                print(f"Chart Attempt {attempt_count} failed with error: {repr(e)}")
                fig1_str = str(fig1) if fig1 is not None else "None"
                fig2_str = str(fig2) if fig2 is not None else "None"
                businessQuestion += f"\nCHART CODE FAILED!  Attempt {attempt_count} failed with error: {repr(e)}\nFig1: {fig1_str}\nFig2: {fig2_str}"

                if attempt_count >= max_attempts:
                    print("Max charting attempts reached, handling the failure.")
                    st.write("I was unable to plot the data.")
                    # Handle the failure after the final attempt
                else:
                    print("Retrying the charts...")

        try:
            with st.expander(label="Business Analysis", expanded=True):
                analysis = analysis_future.result(timeout=30)  # Add a timeout for better handling
                st.markdown(analysis.replace("$", "\$"))
        except:
            st.write("I am unable to provide the analysis. Please rephrase the question and try again.")

    return fig1, fig2, analysis

# Function to create a download link
@st.cache_data(show_spinner=False)
def create_download_link(html_content, filename):
    b64 = base64.b64encode(html_content.encode()).decode()  # B64 encode
    href = f'<a href="data:text/html;base64,{b64}" download="{filename}">Download as HTML</a>'
    return href

@st.cache_data(show_spinner=False)
def read_svg(file_path):
    with open(file_path, 'r') as file:
        content = file.read()
    return content
@st.cache_data(show_spinner=False)
def read_svg_as_base64(file_path):
    with open(file_path, 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')

# Callback function to generate HTML content
@st.cache_data(show_spinner=False)
def generate_html_report(businessQuestion, sqlcode, results, fig1, fig2, analysis, datarobot_logo_svg, customer_logo_svg):
    plotly_html1 = pio.to_html(fig1, full_html=False, include_plotlyjs=True, default_width="100%",
                               default_height="100%")
    plotly_html2 = pio.to_html(fig2, full_html=False, include_plotlyjs=True, default_width="100%",
                               default_height="100%")

    # Convert markdown to HTML for the analysis section
    if analysis and analysis.strip():
        analysis_html = markdown.markdown(analysis)
    else:
        st.error("No analysis data found to generate the report.")

    html_content = f"""
    <html>
    <head>
        <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;700&family=IBM+Plex+Mono:wght@400;700&display=swap">
        <style>
            body {{
                font-family: 'IBM Plex Sans', sans-serif;
                background-color: #F0F1F2;
                color: #0D0D0D;
                margin: 40px;
            }}
            h1, h2 {{
                font-family: 'IBM Plex Sans', sans-serif;
                color: #0D0D0D;
            }}
            pre, code {{
                font-family: 'IBM Plex Mono', monospace;
            }}
            .report-title {{
                font-size: 2.5em;
                font-weight: bold;
                text-align: left;
                margin-top: 40px;
            }}
            .section-title {{
                font-size: 1.75em;
                font-weight: bold;
                margin-top: 20px;
            }}
            .logo-container {{
                text-align: left;
                margin-bottom: 20px;
            }}
            .logo-datarobot {{
                width: 300px;
                margin-bottom: 10px;
                display: block;
            }}
            .logo-customer {{
                width: 300px;
                margin-bottom: 10px;
                display: block;
            }}
            .horizontal-rule {{
                border: 0;
                height: 2px;
                background: #03A688;
                margin: 20px 0;
            }}
            .collapsible {{
                background-color: #03A688;
                color: white;
                cursor: pointer;
                padding: 10px;
                width: 100%;
                border: none;
                text-align: left;
                outline: none;
                font-size: 18px;
                font-weight: bold;
            }}
            .collapsible:after {{
                content: '+';
                font-size: 18px;
                float: right;
            }}
            .collapsible.active:after {{
                content: '-';
            }}
            .content {{
                padding: 0 18px;
                display: none;
                overflow: hidden;
                background-color: #f9f9f9;
            }}
            .content.show {{
                display: block;
            }}
        </style>
        <title>AI Data Analyst Report</title>
    </head>
    <body>
        <div class="logo-container">            
            <img src="data:image/svg+xml;base64,{datarobot_logo_svg}" class="logo-customer" alt="Customer Logo">
        </div>
        <h1 class="report-title">AI Data Analyst Report</h1>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible active">Business Question</button>
        <div class="content show">
            <p>{businessQuestion}</p>
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible">Analysis Code</button>
        <div class="content">
            <pre>{sqlcode}</pre>
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible">Results</button>
        <div class="content">
            {results.to_html(index=False, escape=False)}
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible active">Charts</button>
        <div class="content show">
            <div>{plotly_html1}</div>
            <div>{plotly_html2}</div>
        </div>
        <hr class="horizontal-rule">
        <button type="button" class="collapsible active">Business Analysis</button>
        <div class="content show">
            <div>{analysis_html}</div>
        </div>
        <script>
            var coll = document.getElementsByClassName("collapsible");
            for (var i = 0; i < coll.length; i++) {{
                coll[i].addEventListener("click", function() {{
                    this.classList.toggle("active");
                    var content = this.nextElementSibling;
                    if (content.style.display === "block" || content.classList.contains("show")) {{
                        content.style.display = "none";
                        content.classList.remove("show");
                    }} else {{
                        content.style.display = "block";
                        content.classList.add("show");
                    }}
                }});
            }}
        </script>
    </body>
    </html>
    """
    return html_content

@st.cache_data(show_spinner=False)
def process_tables(dictionary, selectedTables, sampleSize):
    tableSamples = []
    tableDescriptions = []
    frequentValues = pd.DataFrame()

    for table in selectedTables:
        tableDescription = summarizeTable(dictionary, table)
        results = getTableSample(sampleSize=sampleSize, table=table)
        tableSamples.append(results)
        tableDescriptions.append(tableDescription)
        freqVals = get_top_frequent_values(results)
        frequentValues = pd.concat([frequentValues, freqVals], axis=0)

    smallTableSamples = []
    for table in tableSamples:
        smallSample = table.sample(n=3)
        smallTableSamples.append(smallSample)

    return tableDescriptions, tableSamples, smallTableSamples, frequentValues

@st.cache_data(show_spinner=False)
def getSnowflakeTables(user, password, account, database, schema, warehouse):
    # Establish the connection
    conn = snowflake.connector.connect(
        user=user,
        password=password,
        account=account,
        warehouse=warehouse,
        database=database,
        schema=schema,
        # Enable case sensitivity for identifiers
        case_sensitive_identifier_quoting=True
    )

    try:
        # Create a cursor object
        cursor = conn.cursor()

        # Set the warehouse
        cursor.execute(f"USE WAREHOUSE {warehouse}")

        # Execute a query to fetch the table names
        cursor.execute(f"""
                    SELECT table_name
                    FROM information_schema.tables
                    WHERE table_schema = '{schema}'
                """)

        # Fetch all table names
        tables = [row[0] for row in cursor.fetchall()]
        tables.sort()

        return tables

    finally:
        # Close the cursor and connection
        cursor.close()
        conn.close()

# callback functions for the ask button / clear text button
def text_input_enterKey():
    st.session_state["askButton"] = True

def clear_text():
    st.session_state["businessQuestion"] = ""
    st.session_state["askButton"] = False

def make_dictionary_chunks(df):
    dictionary_chunks = []
    chunk_size = 15
    total_columns = len(df.columns)
    progress_placeholder = st.empty()

    for start in range(0, total_columns, chunk_size):
        current_chunk = start // chunk_size + 1
        total_chunks = (total_columns + chunk_size - 1) // chunk_size
        progress = current_chunk / total_chunks

        with progress_placeholder.container():
            st.progress(progress,
                        text=f'Processing {chunk_size} columns at a time in chunks. Currently working on chunk {current_chunk} of {total_chunks}')

        end = min(start + chunk_size, total_columns)
        subset = df.iloc[:10, start:end]
        data = "First 10 Rows: \n" + str(
            subset) + "\n Unique and Frequent Values of Categorical Data: \n" + str(
            get_top_frequent_values(df))

        dictionary_chunk = getDataDictionary(data)
        dictionary_chunks.append(dictionary_chunk)

    progress_placeholder.empty()
    return dictionary_chunks

def render_header():
    st.image("DataRobotLogo.svg", width=300)
    # st.image("small_square_placeholder.svg", width=300)
    st.title("Ask a question about the data.")

def setup_sidebar():
    with st.sidebar:
        st.image("Snowflake.svg", width=75)
        load_snowflake_tables()

        with st.form(key='table_selection_form'):
            st.session_state['selectedTables'] = st.multiselect(
                label="Choose a few tables", options=st.session_state["tables"], key="table_select_box")
            st.session_state["snowflake_submit_button"] = st.form_submit_button(label='Analyze', type="secondary")
        process_table_selection()

        st.image("csv_File_Logo.svg", width=45)
        st.session_state["csvUploadButton"] = st.file_uploader(label="Or, upload a CSV file",
                                                               accept_multiple_files=False)
        process_csv_upload()
        with st.expander("Clear Cache", expanded=False):
            st.write("To reset any saved data and completely start over, clear the cache. You will have to reload your dataset.")
            st.button("Clear Cache", on_click=clear_cache_callback)
            if st.session_state["cache_cleared"]:
                st.success("Cache cleared successfully!")
                # Reset the flag
                st.session_state["cache_cleared"] = False

def load_snowflake_tables():
    try:
        st.session_state["tables"] = getSnowflakeTables(
            user, st.session_state["password"], account, database, schema, warehouse)
    except Exception as e:
        print("Error connecting: ", e)
        st.session_state["tables"] = ["None"]

def process_table_selection():
    if st.session_state["snowflake_submit_button"]:
        st.session_state["table_selection_button"] = True

def process_csv_upload():
    if st.session_state["csvUploadButton"] is not None:
        st.session_state["selectedCSVFile"] = st.session_state["csvUploadButton"]

def display_logo_header():
    st.image("DataRobotLogo.svg", width=300)
    # st.image("small_square_placeholder.svg", width=300)
    st.header("Ask a question about the data.")

def get_data_definitions_and_suggestions():
    with st.spinner("Getting table definitions..."):
        dictionary = getSnowflakeTableDescriptions(
            st.session_state['selectedTables'], user,
            st.session_state["password"], account,
            warehouse, database, schema)

        suggestedQuestions = suggestQuestion(dictionary)
        table_descriptions, table_samples, small_table_samples, frequent_values = process_tables(
            dictionary,
            st.session_state['selectedTables'],
            sampleSize=1000)
        st.session_state.update({
            "tableDescriptions": table_descriptions,
            "tableSamples": table_samples,
            "smallTableSamples": small_table_samples,
            "frequentValues": frequent_values,
        })
    return dictionary, suggestedQuestions

def display_analysis_tab(tab):
    with tab:
        st.write(st.session_state["suggestedQuestions"])

        st.session_state["businessQuestion"] = st.text_input(
            label="Question",
            # value=st.session_state["businessQuestion"],
            on_change=text_input_enterKey
        )
        display_action_buttons()

        if st.session_state.get("askButton", False):
            analyze_question()

def display_explore_tab(tab):
    with tab:
        for i in range(len(st.session_state["tableSamples"])):
            st.subheader(st.session_state['selectedTables'][i])
            st.caption(f"Displaying a random sample of {len(st.session_state['tableSamples'][i])} rows")
            st.write(st.session_state["tableDescriptions"][i])
            st.write(st.session_state["tableSamples"][i])
            display_data_dictionary(i)


def display_data_dictionary(index):
    table_name = st.session_state['selectedTables'][index]
    dictionary_key = f'{table_name}_dictionary'

    if dictionary_key not in st.session_state:
        with st.expander(label=f"Data Dictionary for {table_name}", expanded=False):
            with st.spinner("Making dictionary..."):
                dictionary_chunks = make_dictionary_chunks(st.session_state["tableSamples"][index])
            with st.spinner("Putting it all together..."):
                assembled_dictionary = assembleDictionaryParts(dictionary_chunks)
                st.session_state[dictionary_key] = assembled_dictionary

                # Initialize or append to llm_generated_dictionary
                if 'llm_generated_dictionary' not in st.session_state:
                    st.session_state['llm_generated_dictionary'] = assembled_dictionary
                else:
                    st.session_state['llm_generated_dictionary'] += "\n" + assembled_dictionary

                st.markdown(assembled_dictionary)
    else:
        with st.expander(label=f"Data Dictionary for {table_name}", expanded=False):
            st.markdown(st.session_state[dictionary_key])

def display_csv_explore_tab(tab):
    with tab:
        st.session_state["df"] = pd.read_csv(st.session_state["selectedCSVFile"])
        with st.expander(label="First 10 Rows", expanded=False):
            st.dataframe(st.session_state["df"].head(10))

        try:
            with st.expander(label="Column Descriptions", expanded=False):
                st.dataframe(st.session_state["df"].describe(include='all'))
        except:
            pass

        try:
            with st.expander(label="Unique and Frequent Values", expanded=False):
                st.dataframe(get_top_frequent_values(st.session_state["df"]))
        except Exception as e:
            print(e)

        try:
            with st.expander(label="Data Dictionary", expanded=True):
                with st.spinner("Making dictionary..."):
                    st.session_state['dictionary_chunks'] = make_dictionary_chunks(st.session_state["df"])
                with st.spinner("Putting it all together..."):
                    st.session_state["dictionary"] = assembleDictionaryParts(st.session_state['dictionary_chunks'])
                    st.markdown(st.session_state["dictionary"])
        except:
            pass

def display_csv_analysis_tab(tab):
    with tab:
        st.session_state["suggestedQuestions"] = suggestQuestion(st.session_state["dictionary"])
        st.write(st.session_state["suggestedQuestions"])

        st.session_state["businessQuestion"] = st.text_input(
            label="Question",
            # value=st.session_state["businessQuestion"],
            on_change=text_input_enterKey
        )
        display_action_buttons()

        if st.session_state.get("askButton", False):
            analyze_question_csv()

def display_action_buttons():
    buttonContainer = st.container()
    buttonCol1, buttonCol2, _ = buttonContainer.columns([1, 1, 8])

    buttonCol1.button(label="Ask", use_container_width=True, type="primary", on_click=text_input_enterKey)
    buttonCol2.button(label="clear", use_container_width=True, type="secondary", on_click=clear_text)

def analyze_question():
    with st.spinner("Analyzing... "):
        full_dictionary = []
        st.session_state["prompt"] = generate_prompt()
        execute_query_with_retries(csv_mode=False)

        try:
            display_query_results()
        except:
            st.write(
                "I tried a few different ways, but couldn't get a working solution. Rephrase the question and try again.")

        if st.session_state["results"] is not None and not st.session_state["results"].empty:
            analyze_and_generate_report(full_dictionary)
        else:
            st.write("The query returns an empty result. Try rephrasing the question.")
            print("No data returned.")
            st.stop()

def analyze_question_csv():
    with st.spinner("Analyzing... "):
        st.session_state["prompt"] = generate_csv_prompt()
        execute_query_with_retries(csv_mode=True)

        try:
            display_query_results()
        except:
            st.write(
                "I tried a few different ways, but couldn't get a working solution. Rephrase the question and try again.")

        if st.session_state["results"] is not None and not st.session_state["results"].empty:
            analyze_and_generate_report_csv()
        else:
            st.write("The query returns an empty result. Try rephrasing the question.")
            print("No data returned.")
            st.stop()


def generate_prompt():
    # Ensure the llm_generated_dictionary is not None or empty
    full_dictionary = st.session_state.get('llm_generated_dictionary', '')

    # Build the prompt
    prompt = (
        f"Business Question: {st.session_state.get('businessQuestion', '')}\n"
        f"Data Dictionary: \n{full_dictionary}\n"
        f"Column Definitions: \n{st.session_state.get('tableDescriptions', '')}\n"
        f"Data Sample: \n{st.session_state.get('smallTableSamples', '')}\n"
        f"Frequent Values: \n{st.session_state.get('frequentValues', '')}"
    )

    # Debugging output
    print("\n ================= \n PROMPT \n =================")
    print(prompt)

    return prompt

def generate_csv_prompt():
    return ("Business Question: " + str(st.session_state["businessQuestion"]) +
            "\n Data Sample: \n" + str(st.session_state["df"].head(3)) +
            "\n Unique and Frequent Values of Categorical Data: \n" + str(
                get_top_frequent_values(st.session_state["df"])) +
            "\n Data Dictionary: \n" + str(st.session_state["dictionary"]))

def execute_query_with_retries(csv_mode):
    attempts = 0
    max_retries = 5
    while attempts < max_retries:
        st.session_state["sqlCode"] = None
        try:
            if csv_mode:
                st.session_state["sqlCode"], st.session_state["results"] = executePythonCode(st.session_state["prompt"], st.session_state["df"])
            else:
                st.session_state["sqlCode"], st.session_state["results"] = executeSnowflakeQuery(st.session_state["prompt"], user, st.session_state["password"], account, warehouse, database, schema)
                # st.session_state["sqlCode"], st.session_state["results"] = executeSnowflakeSnowpark(st.session_state["prompt"], user, st.session_state["password"], account, warehouse, database, schema)
            if st.session_state["results"].empty:
                raise ValueError("The DataFrame is empty, retrying...")
            break
        except Exception as e:
            attempts += 1
            st.session_state[
                "prompt"] += f"\nQUERY FAILED! Attempt {attempts} failed with error: {repr(e)}\nCode: {st.session_state['sqlCode']}"
            if attempts == max_retries:
                break

def display_query_results():
    with st.expander(label="Code", expanded=False):
        st.code(st.session_state["sqlCode"], language="sql")
    with st.expander(label="Result", expanded=False):
        st.table(st.session_state["results"])

def analyze_and_generate_report(full_dictionary):
    with st.spinner("Visualization and analysis in progress..."):
        st.session_state["fig1"], st.session_state["fig2"], st.session_state[
            "analysis"] = createChartsAndBusinessAnalysis(
            st.session_state["businessQuestion"],
            st.session_state["results"], st.session_state["prompt"])

    generate_report(full_dictionary)

def analyze_and_generate_report_csv():
    with st.spinner("Visualization and analysis in progress..."):
        st.session_state["fig1"], st.session_state["fig2"], st.session_state[
            "analysis"] = createChartsAndBusinessAnalysis(
            st.session_state["businessQuestion"],
            st.session_state["results"], st.session_state["prompt"])

    generate_report_csv()

def generate_report(full_dictionary):
    read_svgs_and_generate_html_report()
    create_and_display_download_link()
    read_svgs_and_generate_excel_report()
    create_and_display_excel_download_link()

def generate_report_csv():
    read_svgs_and_generate_html_report()
    create_and_display_download_link()
    read_svgs_and_generate_excel_report()
    create_and_display_excel_download_link()


def read_svgs_and_generate_html_report():
    st.session_state["datarobot_logo_svg"] = read_svg_as_base64("DataRobotLogo.svg")
    st.session_state["customer_logo_svg"] = read_svg_as_base64("small_square_placeholder.svg")

    st.session_state["html_content"] = generate_html_report(st.session_state["businessQuestion"],
                                                            st.session_state["sqlCode"],
                                                            st.session_state["results"], st.session_state["fig1"],
                                                            st.session_state["fig2"],
                                                            st.session_state["analysis"],
                                                            st.session_state["datarobot_logo_svg"],
                                                            st.session_state["customer_logo_svg"])
def create_and_display_download_link():
    try:
        st.session_state["download_link"] = create_download_link(st.session_state["html_content"], 'report.html')
        st.markdown(st.session_state["download_link"], unsafe_allow_html=True)
    except:
        pass
@st.cache_data(show_spinner=False)
def create_download_link_excel(excel_data, filename):
    if not excel_data:
        st.error("Excel content is empty. Cannot create a download link.")
        return ""
    b64 = base64.b64encode(excel_data).decode()  # B64 encode
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download as Excel</a>'
    return href

@st.cache_data(show_spinner=False)
def read_svg(file_path):
    with open(file_path, 'r') as file:
        content = file.read()
    return content

@st.cache_data(show_spinner=False)
def read_svg_as_base64(file_path):
    with open(file_path, 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')

# Callback function to generate Excel content
@st.cache_data(show_spinner=False)
def generate_excel_report(businessQuestion, sqlcode, results, fig1, fig2, analysis):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    try:
        # Writing data to Excel
        if businessQuestion:
            df_business_question = pd.DataFrame({'Business Question': [businessQuestion]})
            df_business_question.to_excel(writer, index=False, sheet_name='Business Question')
            worksheet = writer.sheets['Business Question']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        if results is not None and not results.empty:
            results.to_excel(writer, index=False, sheet_name='Results')
            worksheet = writer.sheets['Results']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
            worksheet.auto_filter.ref = worksheet.dimensions

        # Writing analysis to Excel
        if analysis and analysis.strip():
            analysis_sections = {
                "The Bottom Line": "",
                "Additional Insights": "",
                "Follow Up Questions": ""
            }
            current_section = None
            for line in analysis.splitlines():
                line = line.strip()
                if line.startswith("###"):
                    header = line.replace("###", "").strip()
                    if header in analysis_sections:
                        current_section = header
                elif current_section:
                    analysis_sections[current_section] += line + "\n"

            # Write each section to individual cells
            worksheet = writer.book.create_sheet(title='Analysis')
            row = 1
            for section, content in analysis_sections.items():
                worksheet[f'A{row}'] = section
                worksheet[f'A{row}'].font = Font(bold=True)
                cell = worksheet[f'A{row + 1}']
                # content = content.strip().replace('', '')  # Replace newlines for better formatting
                cell.value = content
                cell.alignment = Alignment(wrap_text=True)
                worksheet.column_dimensions['A'].width = 50
                row += 3

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        else:
            st.error("No analysis data found to generate the report.")

        # Add Plotly charts as images
        worksheet = writer.book.create_sheet(title="Charts")
        if fig1:
            fig1_bytes = fig1.to_image(format="png")
            img1 = Image(BytesIO(fig1_bytes))
            worksheet.add_image(img1, "A1")
        if fig2:
            fig2_bytes = fig2.to_image(format="png")
            img2 = Image(BytesIO(fig2_bytes))
            worksheet.add_image(img2, "A20")

        writer._save()  # Close the writer before accessing the value
        return output.getvalue()
    except Exception as e:
        st.error(f"An error occurred while generating the Excel report: {e}")
        return None

def read_svgs_and_generate_excel_report():
    st.session_state["datarobot_logo_svg"] = read_svg_as_base64("DataRobotLogo.svg")
    st.session_state["customer_logo_svg"] = read_svg_as_base64("small_square_placeholder.svg")

    st.session_state["excel_content"] = generate_excel_report(st.session_state.get("businessQuestion"),
                                                               st.session_state.get("sqlCode"),
                                                               st.session_state.get("results"), st.session_state.get("fig1"),
                                                               st.session_state.get("fig2"),
                                                               st.session_state.get("analysis"))

def create_and_display_excel_download_link():
    st.session_state["download_link_excel"] = create_download_link_excel(st.session_state.get("excel_content"), 'report.xlsx')
    if st.session_state["download_link_excel"]:
        st.markdown(st.session_state["download_link_excel"], unsafe_allow_html=True)




def clear_cache_callback():
    # Clear both data and resource caches
    st.cache_data.clear()
    st.cache_resource.clear()

    # Update session state to show success message
    st.session_state["cache_cleared"] = True

def mainPage():
    setup_sidebar()

    display_logo_header()

   if st.session_state["table_selection_button"] or st.session_state["selectedCSVFile"]:
       tab1, tab2 = st.tabs(["Analyze", "Explore"])
       if st.session_state.get("table_selection_button", False):
           st.session_state["dictionary"], st.session_state["suggestedQuestions"] = get_data_definitions_and_suggestions()
           with st.spinner(text="Analyzing table structure, see Explore tab for details..."):
               display_explore_tab(tab2)
           display_analysis_tab(tab1)

       if st.session_state.get("selectedCSVFile", None) is not None:
           with st.spinner(text="Analyzing table structure, see Explore tab for details..."):
               display_csv_explore_tab(tab2)
           display_csv_analysis_tab(tab1)


# Main app
def _main():
    hide_streamlit_style = """
    <style>
    # MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """
    st.markdown(hide_streamlit_style, unsafe_allow_html=True)  # This lets you hide the Streamlit branding

    mainPage()


if __name__ == "__main__":
    _main()
